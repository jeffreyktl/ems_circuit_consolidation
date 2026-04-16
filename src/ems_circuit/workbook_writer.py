from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import CircuitRegisterRow, ConflictRecord, EvidenceRow, PageRecord, SourceFileRecord

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_GRAY_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))
REVIEW_FILL = PatternFill("solid", fgColor="FCE4D6")


def write_workbook(
    output_path: Path,
    school_code: str,
    files: list[SourceFileRecord],
    pages: list[PageRecord],
    evidence_rows: list[EvidenceRow],
    circuit_rows: list[CircuitRegisterRow],
    conflicts: list[ConflictRecord],
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet("Summary")
    _write_sheet(summary_ws, [
        {"metric": "School Code", "value": school_code},
        {"metric": "PDF files detected", "value": len(files)},
        {"metric": "Pages processed", "value": len(pages)},
        {"metric": "Evidence rows", "value": len(evidence_rows)},
        {"metric": "Circuit register rows", "value": len(circuit_rows)},
        {"metric": "Conflicts / review issues", "value": len(conflicts)},
        {"metric": "Distribution board test result pages", "value": sum(1 for p in pages if p.page_subtype == "distribution_board_test_result")},
        {"metric": "Schematic pages", "value": sum(1 for p in pages if p.page_subtype == "schematic_as_built")},
    ])

    _write_sheet(wb.create_sheet("Circuit_Register"), [row.to_dict() for row in circuit_rows], freeze_panes="A2")
    _write_sheet(wb.create_sheet("Source_Evidence"), [row.to_dict() for row in evidence_rows], freeze_panes="A2")
    _write_sheet(wb.create_sheet("Conflicts_Review"), [row.to_dict() for row in conflicts], freeze_panes="A2")
    _write_sheet(wb.create_sheet("Missing_Documents"), _missing_document_rows(pages), freeze_panes="A2")
    _write_sheet(wb.create_sheet("Page_Index"), [row.to_dict() for row in pages], freeze_panes="A2")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _write_sheet(ws, rows: list[dict], freeze_panes: str | None = None) -> None:
    if not rows:
        ws["A1"] = "No records"
        return

    headers = list(rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_GRAY_BORDER

    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        values = [str(cell.value or "") for cell in row]
        if any(token in " | ".join(values) for token in ["Require manual review", "Uncertain", "Yes"]):
            # Only apply review fill to rows obviously needing attention.
            if any(token in " | ".join(values) for token in ["Require manual review", "Uncertain"]):
                for cell in row:
                    cell.fill = REVIEW_FILL

    if freeze_panes:
        ws.freeze_panes = freeze_panes

    _auto_fit_columns(ws)


def _auto_fit_columns(ws) -> None:
    for idx, column in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 80))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_length + 2, 12), 36)


def _missing_document_rows(pages: list[PageRecord]) -> list[dict]:
    subtypes = {page.page_subtype for page in pages}
    return [
        {"document_group": "WR2 / certificate", "status": "Present" if "wr2_certificate" in subtypes else "Missing"},
        {"document_group": "Distribution board test result schedule", "status": "Present" if "distribution_board_test_result" in subtypes else "Missing"},
        {"document_group": "Schematic / as-built drawing", "status": "Present" if "schematic_as_built" in subtypes else "Missing"},
        {"document_group": "Works order", "status": "Present" if "works_order" in subtypes else "Missing"},
        {"document_group": "Transmittal / cover memo", "status": "Present" if "transmittal_cover" in subtypes else "Missing"},
    ]
